import openpyxl
# # Create a new workbook and select the active worksheet
# workbook = openpyxl.Workbook()
# sheet = workbook.active

# # Sample data to write to the Excel file
# headers = ['ID', 'First Name', 'Last Name']
# data = [
#     [1, 'Alice', 'Brown'],
#     [2, 'Bob', 'Smith'],
#     [3, 'Charlie', 'Johnson'],
#     [4, 'David', 'Williams'],
#     [5, 'Eva', 'Jones'],
# ]

# # Write the header
# for col_num, header in enumerate(headers, 1):
#     col_letter = openpyxl.utils.get_column_letter(col_num)
#     sheet[f"{col_letter}1"] = header

# # Write the data rows
# for row_num, row in enumerate(data, 2):  # Start from the second row since first is header
#     for col_num, cell_value in enumerate(row, 1):
#         col_letter = openpyxl.utils.get_column_letter(col_num)
#         sheet[f"{col_letter}{row_num}"] = cell_value

# # Save the workbook to a file
# file_path = "sample_file.xlsx"
# workbook.save(file_path)

# print(f"Excel file saved as {file_path}")

lista_med_data = []
def add_row_data_to_sheet():
    tal = input('Tal')
    namn = input('Namn')
    efternamn = input('Efternamn')
    return [tal, namn, efternamn]


while True:
    meny = input(">>>")
    if meny == 'q': break
    
    lista_med_data.append(add_row_data_to_sheet())

# Load the existing Excel file
file_path = "sample_file.xlsx"
workbook = openpyxl.load_workbook(file_path)

# Select the active worksheet or you can select a particular sheet using workbook["Sheet1"]
sheet = workbook.active

# Data to be appended
rows = [
    [222, 'John', 'Doe'],
    [798, 'Jane', 'Smith'],
    [2341234, 'Robert', 'Johnson'],
]

# Append the data to the sheet
for row in rows:
    sheet.append(row)

# Save the updated workbook
workbook.save(file_path)
